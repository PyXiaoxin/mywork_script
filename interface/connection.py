#!/usr/bin/python3
# -*- coding: utf-8 -*-

import os
import re
import csv
import paramiko
import time
from telnetlib import Telnet
import logging
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from concurrent import futures
from .public_env import get_value


#   变量说明
#   ip = ip地址
#   username = 用户名
#   password = 密码
#   port = 端口号
#   by xiaoxin
# 免责声明：仅用于自己测试,出了问题，概不负责


class deviceControl:  # 交换机登陆模块
    def __init__(self, ip, username, password, port=22):
        self.password = password
        self.username = username
        self.ip = ip
        self.port = port

    def connectDevice(self):  # 适用于连接路由，交换机。登录成功返回True
        times = 0
        # paramiko.util.log_to_file('paramiko.log')  # 调试日志
        while True:  # 尝试3次登陆
            try:
                self.ssh = paramiko.SSHClient()
                policy = paramiko.AutoAddPolicy()
                self.ssh.set_missing_host_key_policy(policy)
                self.ssh.connect(self.ip, self.port, self.username, self.password, timeout=200)
                self.ssh_shell = self.ssh.invoke_shell()  # 使用invoke是为了可以执行多条命令
                self.ssh_shell.settimeout(1)  # tunnel超时
                return True
            except:
                time.sleep(1)
                times += 1
                if times == 2:  # 超时次数=3返回错误
                    self.close()  # 关闭会话
                    return False

    def sendCmd(self, cmd):  # 发送命令(PS:加上了回车符)，返回发送的字节数
        _cmd = cmd
        status = self.ssh_shell.send(' %s\n' % _cmd)
        return status

    def recData(self):  # 接受返回数据
        dataAll = ''  # 定义一个变量用于返回数据
        par = re.compile(r'---- More ----')
        while True:
            data = ''
            times = 0  # 循环次数叠加
            while times <= 3:  # 取一次数据
                try:
                    rec = self.ssh_shell.recv(1024)
                    if bool(rec) is False:
                        raise ValueError
                    data += rec.decode('utf-8')
                    times = 0
                except:
                    times += 1
            if bool(data) is False:  # 获取的数据为空则跳出循环
                break
            else:  # 判断是否有more关键字
                endMark = par.search(data)
                if endMark:
                    self.ssh_shell.send(' ')
                dataAll += data
        return deleteUnknownStr(dataAll)

    def close(self):  # 关闭session
        try:
            self.ssh_shell.close()
        except:
            pass
        try:
            self.ssh.close()
        except:
            pass

    def connectLinux(self):  # 适用于连接F5/netscaler设备
        try:
            self.t = paramiko.Transport(sock=(self.ip, 22))
            self.t.connect(username=self.username, password=self.password)
            self.chan = self.t.open_session(timeout=5)
            self.chan.settimeout(0.5)  # 设置session超时
            self.chan.invoke_shell()
            return True
        except:
            self.close()
            return False

    def sendCmdLinux(self, cmd):  # 适用于F5/netscaler设备发送命令（PS:自带回车符），返回运行结果
        cmd += '\n'  # 命令加上回车符
        result = ''
        self.chan.send(cmd)  # 发送要执行的命令
        times = 1
        while True:  # 回显很长的命令可能执行较久，通过循环10次获取回显信息
            time.sleep(0.5)
            try:
                ret = self.chan.recv(10240)
                result += ret.decode('utf-8')
            except:
                times += 1
                if times == 5:
                    break
        return result

    def telnetConnect(self):
        times = 0
        while True:
            try:
                self.tn = Telnet(self.ip, port=23, timeout=10)
                break
            except:
                times += 1
                if times == 3:
                    self.telnetClose()
                    return False
        # 输入登录用户名
        self.tn.read_until(b'Username:', timeout=10)
        self.tn.write(self.username.encode('ascii') + b'\n')
        # 输入登录密码
        self.tn.read_until(b'Password:', timeout=10)
        self.tn.write(self.password.encode('ascii') + b'\n')
        times = 0
        while True:
            time.sleep(0.5)
            loginInfo = self.tn.read_very_eager()
            # print(times,loginInfo)
            if loginInfo.endswith(b'>'):  # 判断是否登录成功
                return True
            else:
                times += 1
                if times == 3:  # 尝试4次不成功则返回错误
                    self.telnetClose()
                    return False

    def telnetSendReturn(self, cmd):  # 发送命令并获取数据
        _cmd = cmd
        data = ''  # 定义一个变量获取返回的
        times = 0
        try:
            _cmd = _cmd.encode('ascii') + b'\n'
            self.tn.write(_cmd)
        except:
            return data
        while True:
            time.sleep(0.5)
            try:
                rec = self.tn.read_very_eager()
                rec = rec.decode('utf-8')
            except:
                rec = self.tn.read_very_eager()
            data += str(rec)
            if data.endswith(('---- More ----')):
                self.tn.write(' '.encode('ascii'))
            else:
                times += 1
                if times == 5:
                    break
        # data = re.re('  ---- More ----\x1b[42D                                          \x1b[42D', '')
        data = deleteUnknownStr(data)
        return data

    def telnetClose(self):  # 关闭连接
        try:
            self.tn.close()
        except:
            pass


class deviceControl_auto(deviceControl):  # 继承deviceControl的简洁登录 SSH TELNET合并
    def __init__(self, ip, username, password, port=22):  # 继承构造方法
        deviceControl.__init__(self, ip, username, password, port)

    def sendCmd_auto(self, cmd_list: list):  # 使用Telnet SSH 执行多条命令返回结果
        cmd_local = list(set(cmd_list))  # 去重复list
        result = {}  # 命令返回的结果
        ssh_login = deviceControl.connectDevice(self)  # 使用父类SSH登录
        if ssh_login:
            loginWay = 'SSH'
            deviceControl.recData(self)  # 欢迎数据获取
            for cmd in cmd_local:
                deviceControl.sendCmd(self, cmd)
                rec_data = deviceControl.recData(self)
                result[cmd] = rec_data
                result['loginWay'] = loginWay
            deviceControl.close(self)  # 关闭会话
        else:
            telnet_login = deviceControl.telnetConnect(self)
            if telnet_login:
                loginWay = 'TELNET'
                for cmd in cmd_local:
                    rec_data = deviceControl.telnetSendReturn(self, cmd)
                    result[cmd] = rec_data
                    result['loginWay'] = loginWay
                deviceControl.telnetClose(self)
            else:
                raise RuntimeError('SSH&TELNET CONNECT ERROR')
        return result


def deleteUnknownStr(line_p):  # 删除垃圾字符，转义序列字符
    line, i, imax = '', 0, len(line_p)
    while i < imax:
        ac = ord(line_p[i])
        if (32 <= ac < 127) or ac in (9, 10):  # printable, \t, \n
            line += line_p[i]
        elif ac == 27:  # remove coded sequences
            i += 1
            while i < imax and line_p[i].lower() not in 'abcdhsujkm':
                i += 1
        elif ac == 8 or (ac == 13 and line and line[-1] == ' '):  # backspace or EOL spacing
            if line:
                line = line[:-1]
        i += 1
    line = re.sub('\s{2}---- More ----\s+', '', line)
    return line


class excel:  # Excel表格处理 只支持.xlsx格式
    def __init__(self, filename):  # 初始化
        self.filename = filename  # 文件名 .xlsx

    # 写入数据
    def excel_write(self, title, data, sheetname='data01', sheetIndex=1):  # 一次性写入 data 格式为[[],[]]
        self.wb_obj = Workbook()
        self.wb_obj.active
        title_local = title  # 标题 list
        data_local = data
        sheetname_local = sheetname  # sheet名称
        sheetIndex_local = sheetIndex - 1  # sheet的位置 默认是第一张表 位置从0开始
        wsObj = self.wb_obj.create_sheet(sheetname_local, sheetIndex_local)
        wsObj.append(title_local)
        title_font = Font(b='bold', size='12')
        for i in range(1, len(title) + 1):
            cell = wsObj.cell(row=1, column=i)
            cell.font = title_font
        for row_data in data_local:
            try:
                wsObj.append(row_data)  # 写入数据
            except:
                for row in row_data:
                    wsObj.append(row)

    def excel_creat(self, title, sheetname='data01', sheetIndex=1):  # 创建对象并设置好列头
        self.wb_obj = Workbook()
        self.wb_obj.active
        title_local = title  # 标题 list
        sheetname_local = sheetname  # sheet名称
        sheetIndex_local = sheetIndex - 1  # sheet的位置 默认是第一张表 位置从0开始
        self.wsobj = self.wb_obj.create_sheet(sheetname_local, sheetIndex_local)
        self.wsobj.append(title_local)
        title_font = Font(b='bold', size='12')
        for i in range(1, len(title) + 1):
            cell = self.wsobj.cell(row=1, column=i)
            cell.font = title_font

    def write_row(self, data: list):  # 写入单行数据[]
        row_data = data
        self.wsobj.append(row_data)

    # 保存文件
    def save_file(self):  # 保存文件
        timeNow = time.strftime('%Y-%m-%d_%H%M%S', time.localtime(time.time()))
        filename_save = self.filename.replace('.xlsx', '')
        filename_save = '%s_%s.xlsx' % (filename_save, timeNow)  # 返回文件名
        self.wb_obj.save(filename_save)  # 存盘
        self.wb_obj.close()  # 关闭
        return filename_save

    # 读取数据 默认打开第一个sheet从第二行读
    def excel_read(self, sheetnum=1, row=0, column=0, row_start=2, column_start=1):
        file_local = self.filename
        wb = load_workbook(filename=file_local)  # 打开一个excel对象
        sheetnames = wb.sheetnames  # 获取sheets
        ws = wb[sheetnames[sheetnum - 1]]  # 打开第X个sheet
        row_start_local = row_start  # 起始行
        column_start_local = column_start  # 起始列
        if row == 0:
            row = ws.max_row  # 最大行
        if column == 0:
            column = ws.max_column  # 最大列
        data_result = []  # 结果集
        for rx in range(row_start_local, row + 1):  # 循环读取每一行sheet数据
            info = []  # 一行的数据集
            for cx in range(column_start_local, column + 1):  # 循环读取每一行的列数据
                cell_info = ws.cell(row=rx, column=cx).value
                info.append(cell_info)
            data_result.append(info)
        wb.close()  # 关闭
        return data_result

    def excelReadCread(self, ):  # 打开excel对象
        file_local = self.filename
        self.wb = load_workbook(filename=file_local)  # 打开一个excel对象

    def excelReadSheet(self, sheetnum=1, row=0, column=0, row_start=2, column_start=1):  # 读取一个sheet
        sheetnames = self.wb.sheetnames  # 获取sheets
        ws = self.wb[sheetnames[sheetnum - 1]]  # 打开第X个sheet
        row_start_local = row_start  # 起始行
        column_start_local = column_start  # 起始列
        if row == 0:
            row = ws.max_row  # 行计数
        if column == 0:
            column = ws.max_column  # 列计数
        data_result = []  # 结果集
        for rx in range(row_start_local, row + 1):  # 循环读取每一行sheet数据
            info = []  # 一行的数据集
            for cx in range(column_start_local, column + 1):  # 循环读取每一行的列数据
                cell_info = ws.cell(row=rx, column=cx).value
                info.append(cell_info)
            data_result.append(info)
        return data_result

    def excelClose(self):
        self.wb.close()  # 关闭


def readTxt(filename):  # 读取TXT 返回list 忽略#号
    readReturn = []
    with open(filename, 'r', encoding='utf-8') as openfiles:
        readInfo = openfiles.readlines()
    for read_row in readInfo:  # #号行忽略
        readTemp = read_row.strip().startswith('#')
        if readTemp:
            continue
        else:
            readReturn.append(read_row.strip().strip('\n\r'))
    return readReturn


def readCsv(filename):  # 读取CSV文件返回list
    result = []
    with open(filename, mode="r") as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            result.append(row)
    return result


def makeDir(dirName):  # 在当前目录创建文件夹
    path = os.getcwd()  # 获取当前路径
    dir = os.listdir(path)  # 获取当前路径所有文件名
    if dirName not in dir:  # 判断本地是否已经有此文件有则不创建。
        os.mkdir(dirName)


class logg:  # 日志模块
    def __init__(self, loggername, filename):
        # 创建一个logger
        self.logger = logging.getLogger(loggername)
        self.logger.setLevel(logging.DEBUG)

        # 创建一个handler，用于写入日志文件
        timeNow = time.strftime('%Y-%m-%d_%H%M%S', time.localtime(time.time()))
        filename_local = filename
        logname = '%s_%s.log' % (filename_local, timeNow)  # 日志名+日期= 存盘的文件名
        fh = logging.FileHandler(logname, encoding='utf-8')  # 指定utf-8格式编码，避免输出的日志文本乱码
        fh.setLevel(logging.DEBUG)

        # 创建一个handler，用于将日志输出到控制台
        ch = logging.StreamHandler()
        ch.setLevel(logging.ERROR)

        # 定义handler的输出格式
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s:%(message)s',
                                      datefmt='%Y-%m-%d %H:%M:%S')
        fh.setFormatter(formatter)
        ch.setFormatter(formatter)
        # 给logger添加handler
        self.logger.addHandler(fh)
        self.logger.addHandler(ch)

    def get_log(self):
        # 回调logger实例
        return self.logger


class autoThreadingPool():  # 线程池
    def __init__(self, worker=30):
        self.bar = get_value('bar')
        self.worker_local = worker
        self.result = []

    def __call__(self, func, datalist: list):  # 函数，迭代数据
        func_local = func  # function
        datalist_local = datalist  # data
        with futures.ThreadPoolExecutor(max_workers=self.worker_local) as executor:  # max_workers 线程池的数量
            future_list = []
            for row in datalist_local:
                future = executor.submit(func_local, row)
                future_list.append(future)
            unit = 0.8 / len(future_list)
            num = 0.1
            for future in futures.as_completed(future_list):
                res = future.result()
                self.result.append(res)
                num += unit
                self.bar(num)
        return self.result


if __name__ == '__main__':
    # ip = 'XXXX'
    # ip2 = 'XXXX'
    # user = 'XXX'
    # passwd = 'xxxxx'
    # cmds = ['dis clock', 'dis version']
    # conn = deviceContrl_auto(ip, user, passwd)
    # res = conn.sendCmd_auto(cmds)
    # print(res)
    readTxt('../read/keyWords.txt')
